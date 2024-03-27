import requests
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.workbook import Workbook
from domain.utils.constants import Constants
from infraestructure.driven_adapters.automation_metrics_implementation import ElasticSearchAutomationMetrics


def check_and_color_column(index, column, value, condition, columns_to_color):
    if value == condition:
        columns_to_color.extend([(index + 1, column, 'S', 'A')])


def set_column_width(sheet, columns, width):
    for column in columns:
        sheet.column_dimensions[column].width = width


def extract_data(url_and_paths):
    for url, ruta_data in url_and_paths:
        payload = {
            "size": 10000,
            "query": {
                "bool": {
                    "must": [
                        {"term": {"Sprint.keyword": "Sprint 187"}}
                    ],
                    "should": [
                        {"term": {"check_list_perf.keyword": "Se omitio el diligenciamiento y evaluacion del checklist performance"}},
                        {"term": {"check_list_sec.keyword": "Se omitio el diligenciamiento y evaluacion del checklist seguridad"}},
                        {"term": {"titulo_test_plan.keyword": "Titulo no valido o estandar de nombramiento invalido"}},
                        {"term": {"exite_tag.keyword": "Tag no Valido o Estandar Invalido"}},
                        {"term": {"alcance_estrategia.keyword": "No existe o falta informacion en la descripcion(Alcance, estrategia, supuestos, etc)"}},
                        {"term": {"existe_matriz.keyword": "No existe archivo 'Herramienta matriz de riesgos.xlsx' o estandar de nombramiento invalido."}},
                        {"term": {"existe_evidence.keyword": "No existe archivo 'Evidencias_EVCXXX.pdf' o estandar de nombramiento invalido."}}
                    ],
                    "minimum_should_match": 1
                }
            }
        }

        ElasticSearchAutomationMetrics(
            url,
            Constants.Username,
            Constants.Password
        )

        try:
            session = requests.Session()
            session.auth = (Constants.Username, Constants.Password)
            session.verify = False

            response = session.get(url, json=payload)
            print(f"Código de estado de la respuesta: {response.status_code}")

            if response.status_code == 200:
                data = response.json()
                df = pd.json_normalize(data['hits']['hits'])
                df.rename(columns={'_source.Sprint': 'Sprint',
                                   '_source.state': 'Estado',
                                   '_source.dod': 'DOD',
                                   '_source.name_pipeline_pdn': 'Pipeline PDN',
                                   '_source.herramienta_despliegue': 'Herramienta Despliegue',
                                   '_source.evidence_item': 'Id Evidencias VSTS',
                                   '_source.name_pipeline_release_QA': 'Pipeline_Release',
                                   '_source.id_test_plan': 'Testplan',
                                   '_source.stage_E2E': 'Stage E2E',
                                   '_source.stage_manual': 'Stage Manual Test',
                                   '_source.stage_exploratory': 'Stage Exploratory Test',
                                   '_source.existe_matriz': 'Herramienta matriz de riesgos',
                                   '_source.existe_evidence': 'Evidencias',
                                   '_source.check_list_perf': 'Checklist Performance',
                                   '_source.check_list_sec': 'Checklist Seguridad',
                                   '_source.analista': 'Responsable',
                                   '_source.titulo_test_plan': 'Titulo Testplan',
                                   '_source.exite_tag': 'TAG',
                                   '_source.alcance_estrategia': 'Alcance/Estrategia',
                                   '_source.fabrica': 'Fabrica',
                                   '_source.componente_menor': 'Componente Menor',
                                   '_source.lider_componente': 'Lider Inmediato',
                                   }, inplace=True)
                df = df[['Sprint', 'Herramienta Despliegue', 'Estado', 'DOD', 'Pipeline PDN',
                         'Pipeline_Release', 'Id Evidencias VSTS', 'Stage E2E', 'Stage Manual Test',
                         'Stage Exploratory Test', 'Testplan', 'Checklist Performance', 'Checklist Seguridad',
                         'Titulo Testplan', 'TAG', 'Alcance/Estrategia', 'Herramienta matriz de riesgos',
                         'Evidencias', 'Responsable', 'Lider Inmediato', 'Fabrica', 'Componente Menor']]

                df.to_excel(ruta_data, index=False)

                workbook = Workbook()
                sheet = workbook.active
                workbook = pd.ExcelWriter(ruta_data, engine='openpyxl')
                workbook.book = workbook.book
                df.to_excel(workbook, index=False, sheet_name='Sheet1')
                sheet = workbook.sheets['Sheet1']

                for cell in sheet[1]:
                    cell.fill = PatternFill(start_color="FDDA24", end_color="FFFF00", fill_type="solid")

                columns_to_color = []

                for index, row in df.iterrows():
                    checklist_performance_value = row['Checklist Performance']
                    checklist_segue_value = row['Checklist Seguridad']
                    archivo_evidence_value = row['Evidencias']
                    title_testplan_value = row['Titulo Testplan']
                    tag_value = row['TAG']
                    strategy_value = row['Alcance/Estrategia']
                    risk_matrix_value = row['Herramienta matriz de riesgos']

                    check_and_color_column(index, 'L', checklist_performance_value, 'Se omitio el diligenciamiento y evaluacion del checklist performance', columns_to_color)
                    check_and_color_column(index, 'M', checklist_segue_value, 'Se omitio el diligenciamiento y evaluacion del checklist seguridad', columns_to_color)
                    check_and_color_column(index, 'N', title_testplan_value, 'Titulo no valido o estandar de nombramiento invalido', columns_to_color)
                    check_and_color_column(index, 'O', tag_value, 'Tag no Valido o Estandar Invalido', columns_to_color)
                    check_and_color_column(index, 'P', strategy_value, 'No existe o falta informacion en la descripcion(Alcance, estrategia, supuestos, etc)', columns_to_color)
                    check_and_color_column(index, 'Q', risk_matrix_value, "No existe archivo 'Herramienta matriz de riesgos.xlsx' o estandar de nombramiento invalido.", columns_to_color)
                    check_and_color_column(index, 'R', archivo_evidence_value, "No existe archivo 'Evidencias_EVCXXX.pdf' o estandar de nombramiento invalido.", columns_to_color)

                for column in sheet.columns:
                    sheet.column_dimensions[column[0].column_letter].width = 35

                set_column_width(sheet, ['A', 'C', 'D', 'H', 'I', 'J', 'K'], 10)
                set_column_width(sheet, ['B'], 20)

                for row_index, column, column_a, column_b in columns_to_color:
                    cell = sheet[column][row_index]
                    cell_a = sheet[column_a][row_index]
                    cell_b = sheet[column_b][row_index]
                    cell.fill = PatternFill(start_color="cd3f3f", end_color="cd3f3f", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                    cell_a.font = Font(color="A43232")
                    cell_b.font = Font(color="A43232")

                    border = Border(left=Side(style='thin', color='FFFFFF'),
                                right=Side(style='thin', color='FFFFFF'),
                                top=Side(style='thin', color='FFFFFF'),
                                bottom=Side(style='thin', color='FFFFFF'))

                    cell.border = border

                workbook.save()

                print(f"Datos exportados a {ruta_data} exitosamente.")
            else:
                print(f"Fallo al obtener datos. Código de estado: {response.status_code}")
                print(f"Respuesta de la API: {response.text}")

        except Exception as e:
            print(f"Error al obtener datos para la URL: {url}: {e}")


if __name__ == "__main__":
    urls_and_paths_to_extract = [
        (Constants.url_auto, Constants.ruta_data_auto),
        (Constants.url_auto_cer, Constants.ruta_data_auto_cer),
    ]
    extract_data(urls_and_paths_to_extract)
