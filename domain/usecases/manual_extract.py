import json
import re

import requests
import pandas as pd
import row
import os
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.workbook import Workbook
from domain.utils.constants import Constants
from infraestructure.driven_adapters.automation_metrics_implementation import ElasticSearchAutomationMetrics


def check_and_color_column(index, column, value, condition, columns_to_color):
    if value == condition:
        columns_to_color.extend([(index + 1, column, 'A', 'N')])

def set_column_width(sheet, columns, width):
    for column in columns:
        sheet.column_dimensions[column].width = width

def extract_data(url_and_paths):
    for url, ruta_data in url_and_paths:
        payload =  {
            "size": 10000,
            "query": {
                "bool": {
                    "must": [
                        {"term": {"Sprint.keyword": "Sprint 188"}}
                    ],
                    "should": [
                        {"term": {"check_list_perf.keyword": "Se omitio el diligenciamiento y evaluacion del checklist performance"}},
                        {"term": {"check_list_sec.keyword": "Se omitio el diligenciamiento y evaluacion del checklist de seguridad"}},
                        {"term": {"comment_sec.keyword": "No cumple con el concepto por parte del analista de performance, el cual debía relacionar según el resultado en el checklist."}},
                        {"term": {"title_test_plan.keyword": "Titulo no valido o estandar de nombramiento invalido"}},
                        {"term": {"existe_tag_repositorio.keyword": "Tag no Valido o Estandar Invalido"}},
                        {"term": {"tiene_alcance_estrategia.keyword": "No existe o falta informacion en la descripcion(Alcance, estrategia, supuestos, etc)"}},
                        {"term": {"matriz_riesgo_test_plan.keyword": "No existe archivo 'Herramienta matriz de riesgos.xlsx' o estandar de nombramiento invalido."}},
                        {"term": {"evidencias_test_plan.keyword": "No existe archivo 'Evidencias_EVCXXX.pdf' o estandar de nombramiento invalido."}}
                    ],
                    "minimum_should_match": 1
                }
            }
        }

        ElasticSearchAutomationMetrics(
            url,
            Constants.Username,
            Constants.Password
        )

        try:
            session = requests.Session()
            session.auth = (Constants.Username, Constants.Password)
            session.verify = False

            response = session.get(url, json=payload)
            print(f"Código de estado de la respuesta: {response.status_code}")


            if response.status_code == 200:
                data = response.json()
                df = pd.json_normalize(data['hits']['hits'])
                df.rename(columns={'_source.Sprint': 'Sprint',
                                   '_source.depliegue_automatico': 'Herramienta Despliegue',
                                   '_source.RegisterId': 'DOD',
                                   '_source.evidencia': 'Id Evidencias VSTS',
                                   '_source.id_test_plan': 'Testplan',
                                   '_source.check_list_perf': 'Checklist Performance',
                                   '_source.check_list_sec': 'Checklist Seguridad',
                                   '_source.comment_sec': 'Comentario',
                                   '_source.title_test_plan': 'Titulo Testplan',
                                   '_source.existe_tag_repositorio': 'TAG',
                                   '_source.tiene_alcance_estrategia': 'Alcance/Estrategia',
                                   '_source.matriz_riesgo_test_plan': 'Herramienta matriz de riesgos',
                                   '_source.evidencias_test_plan': 'Evidencias',
                                   '_source.analista': 'Responsable Testplan',
                                   '_source.lider_componente': 'Lider Inmediato',
                                   '_source.fabrica': 'Fabrica',
                                   '_source.componente_menor': 'Componente Menor',
                                   }, inplace=True)
                df = df[['Sprint', 'Herramienta Despliegue', 'DOD', 'Id Evidencias VSTS', 'Testplan', 'Checklist Performance',
                         'Checklist Seguridad', 'Comentario', 'Titulo Testplan', 'TAG', 'Alcance/Estrategia', 'Herramienta matriz de riesgos',
                         'Evidencias', 'Responsable Testplan', 'Lider Inmediato', 'Fabrica', 'Componente Menor']]

                df.to_excel(ruta_data, index=False)

                workbook = Workbook()
                sheet = workbook.active
                workbook = pd.ExcelWriter(ruta_data, engine='openpyxl')
                workbook.book = workbook.book
                df.to_excel(workbook, index=False, sheet_name='Sheet1')
                sheet = workbook.sheets['Sheet1']

                for cell in sheet[1]:
                    cell.fill = PatternFill(start_color="FDDA24", end_color="FFFF00", fill_type="solid")

                columns_to_color = []

                for index, row in df.iterrows():
                    validate_testplan_in_dod = row['Testplan']
                    checklist_performance_value = row['Checklist Performance']
                    checklist_segue_value = row['Checklist Seguridad']
                    comment_value= row['Comentario']
                    title_testplan_value = row['Titulo Testplan']
                    tag_value = row['TAG']
                    strategy_value = row['Alcance/Estrategia']
                    risk_matrix_value = row['Herramienta matriz de riesgos']
                    archivo_evidence_value = row['Evidencias']


                    check_and_color_column(index, 'E', validate_testplan_in_dod, "No se relaciono el id del testplan o no cumple con estandar de nombramiento 'Evidencias VSTS: 12345'", columns_to_color)
                    check_and_color_column(index, 'F', checklist_performance_value, 'Se omitio el diligenciamiento y evaluacion del checklist performance', columns_to_color)
                    check_and_color_column(index, 'G', checklist_segue_value, 'Se omitio el diligenciamiento y evaluacion del checklist seguridad', columns_to_color)
                    check_and_color_column(index, 'H', comment_value, 'No cumple con el concepto por parte del analista de performance, el cual debía relacionar según el resultado en el checklist.', columns_to_color)
                    check_and_color_column(index, 'I', title_testplan_value, 'Titulo no valido o estandar de nombramiento invalido', columns_to_color)
                    check_and_color_column(index, 'J', tag_value, 'Tag no Valido o Estandar Invalido. Ejm: AW0157001_ADMINFO_LFS Release 1', columns_to_color)
                    check_and_color_column(index, 'K', strategy_value, 'No existe o falta informacion en la descripcion(Alcance, estrategia, supuestos, etc)', columns_to_color)
                    check_and_color_column(index, 'L', risk_matrix_value, "No existe archivo 'Herramienta matriz de riesgos.xlsx' o estandar de nombramiento invalido.", columns_to_color)
                    check_and_color_column(index, 'M', archivo_evidence_value, "No existe archivo 'Evidencias_EVCXXX.pdf' o estandar de nombramiento invalido.", columns_to_color)

                for column in sheet.columns:
                    sheet.column_dimensions[column[0].column_letter].width = 35

                set_column_width(sheet, ['A', 'B', 'C'], 20)

                for row_index, column, column_a, column_b in columns_to_color:
                    cell = sheet[column][row_index]
                    cell_a = sheet[column_a][row_index]
                    cell_b = sheet[column_b][row_index]

                    cell.fill = PatternFill(start_color="cd3f3f", end_color="cd3f3f", fill_type="solid")
                    cell.font = Font(color="FFFFFF")
                    cell_a.font = Font(color="A43232")
                    cell_b.font = Font(color="A43232")


                    border = Border(left=Side(style='thin', color='FFFFFF'),
                                    right=Side(style='thin', color='FFFFFF'),
                                    top=Side(style='thin', color='FFFFFF'),
                                    bottom=Side(style='thin', color='FFFFFF'))

                    cell.border = border

                workbook.save()

                print(f"Datos exportados a {ruta_data} exitosamente.")
            else:
                print(f"Fallo al obtener datos. Código de estado: {response.status_code}")
                print(f"Respuesta de la API: {response.text}")

        except Exception as e:
            print(f"Error al obtener datos para la URL: {url}: {e}")


if __name__ == "__main__":
    urls_and_paths_to_extract = [
        (Constants.url_manual, Constants.ruta_data_manual)
    ]
    extract_data(urls_and_paths_to_extract)